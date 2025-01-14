import json
import tkinter as tk
from tkinter import messagebox, simpledialog
from tkinter.filedialog import askopenfilename, asksaveasfilename
import openpyxl

class Animal:
    def __init__(self, name: str, age: int):
        """Инициализация атрибутов животного."""
        self.name = name
        self.age = age

    def make_sound(self):
        """Метод, который издает звук. Должен быть переопределен в дочерних классах."""
        raise NotImplementedError("Этот метод должен быть переопределен в подклассах")

    def eat(self, food: str):
        """Метод для описания процесса еды."""
        print(f"{self.name} ест {food}.")

    def to_dict(self):
        """Сериализация животного в словарь."""
        return {
            "type": type(self).__name__,
            "name": self.name,
            "age": self.age
        }

    @staticmethod
    def from_dict(data):
        """Десериализация животного из словаря."""
        animal_type = data["type"]
        if animal_type == "Птица":
            return Bird(data["name"], data["age"], data["wing_span"])
        elif animal_type == "Млекопитающее":
            return Mammal(data["name"], data["age"], data["fur_color"])
        elif animal_type == "Рептилия":
            return Reptile(data["name"], data["age"], data["is_venomous"])

class Bird(Animal):
    def __init__(self, name: str, age: int, wing_span: float):
        """Инициализация атрибутов птицы."""
        super().__init__(name, age)
        self.wing_span = wing_span

    def make_sound(self):
        print(f"{self.name} чирикает.")

    def to_dict(self):
        data = super().to_dict()
        data["wing_span"] = self.wing_span
        data["type"] = "Птица"
        return data

class Mammal(Animal):
    def __init__(self, name: str, age: int, fur_color: str):
        """Инициализация атрибутов млекопитающего."""
        super().__init__(name, age)
        self.fur_color = fur_color

    def make_sound(self):
        print(f"{self.name} рычит или мурлычет.")

    def to_dict(self):
        data = super().to_dict()
        data["fur_color"] = self.fur_color
        data["type"] = "Млекопитающее"
        return data

class Reptile(Animal):
    def __init__(self, name: str, age: int, is_venomous: bool):
        """Инициализация атрибутов рептилии."""
        super().__init__(name, age)
        self.is_venomous = is_venomous

    def make_sound(self):
        print(f"{self.name} шипит.")

    def to_dict(self):
        data = super().to_dict()
        data["is_venomous"] = self.is_venomous
        data["type"] = "Рептилия"
        return data

def animal_sound(animals):
    """Вызывает метод make_sound для каждого животного в списке."""
    for animal in animals:
        animal.make_sound()

class Zoo:
    def __init__(self, name: str):
        """Инициализация зоопарка."""
        self.name = name
        self.animals = []
        self.staff = []

    def add_animal(self, animal: Animal):
        """Добавляет животное в зоопарк."""
        self.animals.append(animal)
        print(f"Животное {animal.name} добавлено в зоопарк {self.name}.")

    def add_staff(self, staff_member):
        """Добавляет сотрудника в зоопарк."""
        self.staff.append(staff_member)
        print(f"Сотрудник {staff_member.name} добавлен в зоопарк {self.name}.")

    def list_animals(self):
        """Выводит список всех животных в зоопарке."""
        return [f"- {animal.name} ({type(animal).__name__})" for animal in self.animals]

    def list_staff(self):
        """Выводит список всех сотрудников зоопарка."""
        return [f"- {staff_member.name} ({type(staff_member).__name__})" for staff_member in self.staff]

    def save_to_file(self, filename: str):
        """Сохраняет состояние зоопарка в файл."""
        data = {
            "name": self.name,
            "animals": [animal.to_dict() for animal in self.animals],
            "staff": [staff.name for staff in self.staff]
        }
        with open(filename, "w") as file:
            json.dump(data, file)
        print(f"Состояние зоопарка сохранено в файл {filename}.")

    def save_to_excel(self, filename: str):
        """Сохраняет состояние зоопарка в Excel файл."""
        wb = openpyxl.Workbook()
        animal_sheet = wb.active
        animal_sheet.title = "Животные"
        animal_sheet.append(["Тип", "Имя", "Возраст", "Доп. атрибут"])

        for animal in self.animals:
            if isinstance(animal, Bird):
                extra = f"Размах крыльев: {animal.wing_span}"
            elif isinstance(animal, Mammal):
                extra = f"Цвет шерсти: {animal.fur_color}"
            elif isinstance(animal, Reptile):
                extra = f"Ядовитое: {'Да' if animal.is_venomous else 'Нет'}"
            else:
                extra = ""
            animal_sheet.append([type(animal).__name__, animal.name, animal.age, extra])

        staff_sheet = wb.create_sheet(title="Сотрудники")
        staff_sheet.append(["Имя", "Должность"])
        for staff in self.staff:
            staff_sheet.append([staff.name, type(staff).__name__])

        wb.save(filename)
        print(f"Состояние зоопарка сохранено в Excel файл {filename}.")

    @staticmethod
    def load_from_file(filename: str):
        """Загружает состояние зоопарка из файла."""
        with open(filename, "r") as file:
            data = json.load(file)
        zoo = Zoo(data["name"])
        zoo.animals = [Animal.from_dict(animal_data) for animal_data in data["animals"]]
        zoo.staff = [Staff(name) for name in data["staff"]]
        print(f"Состояние зоопарка загружено из файла {filename}.")
        return zoo

    @staticmethod
    def load_from_excel(filename: str):
        """Загружает состояние зоопарка из Excel файла."""
        wb = openpyxl.load_workbook(filename)
        zoo = Zoo("Загруженный Зоопарк")

        if "Животные" in wb.sheetnames:
            animal_sheet = wb["Животные"]
            for row in animal_sheet.iter_rows(min_row=2, values_only=True):
                animal_type, name, age, extra = row
                age = int(age)
                if animal_type == "Птица":
                    wing_span = float(extra.split(": ")[1])
                    zoo.add_animal(Bird(name, age, wing_span))
                elif animal_type == "Млекопитающее":
                    fur_color = extra.split(": ")[1]
                    zoo.add_animal(Mammal(name, age, fur_color))
                elif animal_type == "Рептилия":
                    is_venomous = extra.split(": ")[1] == "Да"
                    zoo.add_animal(Reptile(name, age, is_venomous))

        if "Сотрудники" in wb.sheetnames:
            staff_sheet = wb["Сотрудники"]
            for row in staff_sheet.iter_rows(min_row=2, values_only=True):
                name, staff_type = row
                if staff_type == "ZooKeeper":
                    zoo.add_staff(ZooKeeper(name))
                elif staff_type == "Veterinarian":
                    zoo.add_staff(Veterinarian(name))

        print(f"Состояние зоопарка загружено из Excel файла {filename}.")
        return zoo

class Staff:
    def __init__(self, name: str):
        """Инициализация общего сотрудника."""
        self.name = name

class ZooKeeper(Staff):
    def feed_animal(self, animal: Animal, food: str):
        """Кормит животное."""
        print(f"{self.name} кормит {animal.name} {food}.")
        animal.eat(food)

class Veterinarian(Staff):
    def heal_animal(self, animal: Animal):
        """Лечит животное."""
        print(f"{self.name} лечит {animal.name}.")

class ZooApp:
    def __init__(self, root, zoo):
        self.root = root
        self.zoo = zoo
        self.root.title("Система управления зоопарком")

        tk.Button(root, text="Добавить животное", command=self.add_animal).pack(pady=5)
        tk.Button(root, text="Добавить сотрудника", command=self.add_staff).pack(pady=5)
        tk.Button(root, text="Список животных", command=self.list_animals).pack(pady=5)
        tk.Button(root, text="Список сотрудников", command=self.list_staff).pack(pady=5)
        tk.Button(root, text="Сохранить в Excel", command=self.save_to_excel).pack(pady=5)
        tk.Button(root, text="Загрузить из Excel", command=self.load_from_excel).pack(pady=5)

    def add_animal(self):
        animal_type = simpledialog.askstring("Тип животного", "Введите тип животного (Птица, Млекопитающее, Рептилия):")
        name = simpledialog.askstring("Имя животного", "Введите имя животного:")
        age = simpledialog.askinteger("Возраст животного", "Введите возраст животного:")

        if animal_type == "Птица":
            wing_span = simpledialog.askfloat("Размах крыльев", "Введите размах крыльев (в метрах):")
            animal = Bird(name, age, wing_span)
        elif animal_type == "Млекопитающее":
            fur_color = simpledialog.askstring("Цвет шерсти", "Введите цвет шерсти:")
            animal = Mammal(name, age, fur_color)
        elif animal_type == "Рептилия":
            is_venomous = messagebox.askyesno("Ядовитое животное", "Животное ядовитое?")
            animal = Reptile(name, age, is_venomous)
        else:
            messagebox.showerror("Ошибка", "Некорректный тип животного!")
            return

        self.zoo.add_animal(animal)

    def add_staff(self):
        staff_type = simpledialog.askstring("Тип сотрудника", "Введите тип сотрудника (ZooKeeper, Veterinarian):")
        name = simpledialog.askstring("Имя сотрудника", "Введите имя сотрудника:")

        if staff_type == "ZooKeeper":
            staff = ZooKeeper(name)
        elif staff_type == "Veterinarian":
            staff = Veterinarian(name)
        else:
            messagebox.showerror("Ошибка", "Некорректный тип сотрудника!")
            return

        self.zoo.add_staff(staff)

    def list_animals(self):
        animals = self.zoo.list_animals()
        messagebox.showinfo("Список животных", "\n".join(animals))

    def list_staff(self):
        staff = self.zoo.list_staff()
        messagebox.showinfo("Список сотрудников", "\n".join(staff))

    def save_to_file(self):
        filename = asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if filename:
            self.zoo.save_to_file(filename)

    def save_to_excel(self):
        filename = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.zoo.save_to_excel(filename)

    def load_from_file(self):
        filename = askopenfilename(filetypes=[("JSON files", "*.json")])
        if filename:
            self.zoo = Zoo.load_from_file(filename)

    def load_from_excel(self):
        filename = askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.zoo = Zoo.load_from_excel(filename)

if __name__ == "__main__":
    root = tk.Tk()
    zoo = Zoo("Мой Зоопарк")
    app = ZooApp(root, zoo)
    root.mainloop()